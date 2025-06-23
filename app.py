from kivy.app import App
from kivy.lang import Builder
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.filechooser import FileChooserListView
from kivy.uix.popup import Popup
import pandas as pd
import re
from rapidfuzz import fuzz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import shutil
import os

KV = '''
<MainScreen>:
    orientation: "vertical"
    padding: 10
    spacing: 10

    Label:
        text: "برنامج مطابقة الأسماء والأقسام"
        font_size: 24
        size_hint_y: None
        height: self.texture_size[1]

    Button:
        text: "اختيار ملف الأسماء"
        size_hint_y: None
        height: 50
        on_release: root.select_names_file()

    Label:
        id: names_file_label
        text: "لم يتم اختيار ملف الأسماء"
        size_hint_y: None
        height: self.texture_size[1]

    Button:
        text: "اختيار ملف قاعدة البيانات"
        size_hint_y: None
        height: 50
        on_release: root.select_database_file()

    Label:
        id: db_file_label
        text: "لم يتم اختيار ملف قاعدة البيانات"
        size_hint_y: None
        height: self.texture_size[1]

    Button:
        text: "بدء المطابقة"
        size_hint_y: None
        height: 60
        on_release: root.start_matching()

    Label:
        id: status_label
        text: ""
        size_hint_y: None
        height: self.texture_size[1]

    Button:
        text: "حفظ النتائج"
        size_hint_y: None
        height: 50
        on_release: root.save_results()
        disabled: True
'''

class MainScreen(BoxLayout):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.names_file = None
        self.db_file = None
        self.results_df = None
        self.output_path = None

    def select_names_file(self):
        content = FileChooserListView(path='.', filters=['*.xlsx'])
        popup = Popup(title="اختر ملف الأسماء", content=content, size_hint=(0.9, 0.9))
        content.bind(on_submit=lambda instance, selection, touch: self._on_names_file_selected(selection, popup))
        popup.open()

    def _on_names_file_selected(self, selection, popup):
        if selection:
            self.names_file = selection[0]
            self.ids.names_file_label.text = f"تم اختيار: {os.path.basename(self.names_file)}"
        popup.dismiss()

    def select_database_file(self):
        content = FileChooserListView(path='.', filters=['*.xlsx'])
        popup = Popup(title="اختر ملف قاعدة البيانات", content=content, size_hint=(0.9, 0.9))
        content.bind(on_submit=lambda instance, selection, touch: self._on_db_file_selected(selection, popup))
        popup.open()

    def _on_db_file_selected(self, selection, popup):
        if selection:
            self.db_file = selection[0]
            self.ids.db_file_label.text = f"تم اختيار: {os.path.basename(self.db_file)}"
        popup.dismiss()

    def normalize_name(self, name):
        if pd.isnull(name):
            return ""
        name = name.strip().replace("ه", "ة").replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
        name = re.sub(r'(عبد)([^\s])', r' ', name)
        return " ".join(name.split()).lower()

    def is_first_three_words_match(self, name1, name2):
        words1 = name1.split()
        words2 = name2.split()
        length = min(len(words1), len(words2), 3)
        return all(words1[i] == words2[i] for i in range(length))

    def start_matching(self):
        if not self.names_file or not self.db_file:
            self.ids.status_label.text = "يرجى اختيار كلا الملفين أولاً."
            return
        self.ids.status_label.text = "جاري المعالجة..."
        try:
            names_df = pd.read_excel(self.names_file)
            database_df = pd.read_excel(self.db_file)

            names_df["normalized_name"] = names_df["اسم الموظف"].apply(self.normalize_name)
            database_df["normalized_name"] = database_df["اسم الموظف"].apply(self.normalize_name)

            database_df = database_df.drop_duplicates(subset=["normalized_name"])
            database_map = database_df.set_index("normalized_name")[["اسم الموظف", "Operator Id", "المدرسة", "الدائرة"]].to_dict(orient="index")

            matched_results = []

            for original_name, normalized_name in zip(names_df["اسم الموظف"], names_df["normalized_name"]):
                best_match = None
                best_score = 0
                for db_name in database_map.keys():
                    score = fuzz.ratio(normalized_name, db_name)
                    if score > best_score:
                        best_score = score
                        best_match = db_name

                match_data = None
                if best_score >= 85 and (self.is_first_three_words_match(normalized_name, best_match) or best_match.startswith(normalized_name)):
                    match_data = database_map[best_match]
                else:
                    for db_name in database_map.keys():
                        if db_name.startswith(normalized_name):
                            match_data = database_map[db_name]
                            best_match = db_name
                            best_score = fuzz.ratio(normalized_name, best_match)
                            break

                if match_data:
                    matched_results.append({
                        "الاسم الأصلي": original_name,
                        "الاسم المطابق": match_data["اسم الموظف"],
                        "Operator Id": match_data.get("Operator Id", ""),
                        "المدرسة": match_data.get("المدرسة", ""),
                        "الدائرة": match_data.get("الدائرة", ""),
                        "نسبة التطابق": f"{round(best_score)}%",
                        "ملاحظة": "✅ تطابق دقيق"
                    })
                else:
                    matched_results.append({
                        "الاسم الأصلي": original_name,
                        "الاسم المطابق": "",
                        "Operator Id": "",
                        "المدرسة": "",
                        "الدائرة": "",
                        "نسبة التطابق": "",
                        "ملاحظة": "❌ لم يتم العثور على تطابق"
                    })

            results_df = pd.DataFrame(matched_results)
            self.results_df = results_df
            self.output_path = "نتائج_المطابقة_kivy.xlsx"
            results_df.to_excel(self.output_path, index=False)

            # تلوين الصفوف في Excel
            wb = load_workbook(self.output_path)
            ws = wb.active

            for column in ws.columns:
                max_length = 0
                col_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = max_length + 2
                ws.column_dimensions[col_letter].width = adjusted_width

            red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                note = row[6].value  # عمود "ملاحظة"
                if note and "❌" in note:
                    for cell in row:
                        cell.fill = red_fill

            wb.save(self.output_path)
            self.ids.status_label.text = "✅ تمت المطابقة بنجاح! يمكنك الآن حفظ النتائج."
            self.ids['save_results'].disabled = False
        except Exception as e:
            self.ids.status_label.text = f"حدث خطأ: {str(e)}"

    def save_results(self):
        if self.results_df is None:
            self.ids.status_label.text = "لا توجد نتائج للحفظ."
            return
        try:
            save_path = os.path.join(os.path.expanduser('~'), 'نتائج_المطابقة_kivy.xlsx')
            shutil.copy(self.output_path, save_path)
            self.ids.status_label.text = f"✅ تم حفظ الملف في: {save_path}"
        except Exception as e:
            self.ids.status_label.text = f"خطأ في الحفظ: {str(e)}"

class MyApp(App):
    def build(self):
        self.title = "برنامج مطابقة الأسماء والأقسام"
        return Builder.load_string(KV)

if __name__ == "__main__":
    MyApp().run()